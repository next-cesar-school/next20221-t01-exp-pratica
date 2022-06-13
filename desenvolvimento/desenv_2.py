import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import openpyxl_image_loader

worksheet = openpyxl.load_workbook('D:\\O QUE DEIXAR\\Cursos Online\\NExT\\Projeto Final NExT - T1\\F1_imagem.xlsx')
sheet = worksheet.active
image_loader = SheetImageLoader(sheet)

for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        if image_loader.image_in(cell):
            image = image_loader.get(cell)
            image.show()